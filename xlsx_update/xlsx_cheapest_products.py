import os
import openpyxl
import psycopg2


def update_cheapest_products_xlsx():
    connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                                  port=5433)
    cursor = connection.cursor()

    cursor.execute(f'''
    SELECT * 
    FROM products 
    WHERE datetime >= CURRENT_DATE - INTERVAL '3 days';
    ''')
    check_data = cursor.fetchall()
    print(len(check_data))
    if len(check_data) == 0:
        print("Нет данных за последние три дня - произведите повторный парсинг")
    else:

        cursor.execute(f'''
        select * from products where price IN
        (select min(price) from products
        group by category) and  datetime >= CURRENT_DATE - INTERVAL '3 days';
        ''')

        rows = cursor.fetchall()
        print(rows)

        if not os.path.exists(f'../data/e_query_results/query_7-cheapest_products.xlsx'):
            with open(f'../data/e_query_results/query_7-cheapest_products.xlsx', 'w'):
                pass

        filepath = f'../data/e_query_results/query_7-cheapest_products.xlsx'
        workbook_create = openpyxl.Workbook()
        workbook_create.save(filepath)

        wb = openpyxl.load_workbook(f'../data/e_query_results/query_7-cheapest_products.xlsx')

        ws = wb.active

        ws.delete_cols(1, 100)
        ws.delete_rows(1, 3000)

        ws["A1"] = "id"
        ws["B1"] = "datetime"
        ws["C1"] = "shop"
        ws["D1"] = "category"
        ws["E1"] = "product_name"
        ws["F1"] = "price"
        ws["G1"] = "volume"
        ws["H1"] = "price_real"

        for row in rows:
            print(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7])

        for i in range(0, len(rows)):
            ws[f"A{i + 2}"] = rows[i][0]
            ws[f"B{i + 2}"], timezone_removal = str(rows[i][1]).split('+')
            ws[f"C{i + 2}"] = rows[i][2]
            ws[f"D{i + 2}"] = rows[i][3]
            ws[f"E{i + 2}"] = rows[i][4]
            ws[f"F{i + 2}"] = rows[i][5]
            ws[f"G{i + 2}"] = rows[i][6]
            ws[f"H{i + 2}"] = rows[i][7]

        wb.save(f'../data/e_query_results/query_7-cheapest_products.xlsx')

    # cursor.close()
    # connection.close()


if __name__ == '__main__':
    update_cheapest_products_xlsx()