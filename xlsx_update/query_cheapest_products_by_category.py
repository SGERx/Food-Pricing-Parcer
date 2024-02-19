import datetime
import os
import sqlite3
import csv

import openpyxl
import pandas as pd
import psycopg2

from datetime import timezone, timedelta, datetime

current_date = datetime.today().strftime('%Y-%m-%d')
current_date_timezoned = current_date + '+03'
dt = datetime.now(timezone.utc)
print(current_date)
print(current_date_timezoned)
print(dt)


def update_cheapest_products_xlsx():
    # Устанавливаем соединение с базой данных
    connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                                  port=5433)
    cursor = connection.cursor()

    # Проверяем данные за сегодняшнюю дату

    cursor.execute(f'''
    select * from products where datetime - '{dt}' < '1 day';
    ''')
    check_data = cursor.fetchall()
    print(len(check_data))
    if len(check_data) == 0:
        print('Нет данных за сегодняшнюю дату, используйте скрипт query_all_data.py для выгрузки всех данных и ручного '
              'просмотра')
    else:
        # Выбираем все данные
        cursor.execute(f'''
        select * from products where price IN
        (select min(price) from products
        group by category) and  datetime - '{dt}' < '1 day';
        ''')

        rows = cursor.fetchall()
        print(rows)

        test_header = ['datetime', 'shop', 'category', 'product_name', 'price', 'volume', 'price_real', 'id']

        # with open(f'../data/e_query_results/cheapest_products.csv', 'w', newline='') as file:
        #     writer = csv.writer(file)
        #     writer.writerow(test_header)
        #     writer.writerows(rows)

        if not os.path.exists(f'../data/e_query_results/query_3-cheapest_products.xlsx'):
            with open(f'../data/e_query_results/query_3-cheapest_products.xlsx', 'w'):
                pass

        # OpenPyXL
        filepath = f'../data/e_query_results/query_3-cheapest_products.xlsx'
        workbook_create = openpyxl.Workbook()
        workbook_create.save(filepath)

        wb = openpyxl.load_workbook(f'../data/e_query_results/query_3-cheapest_products.xlsx')

        # Создайте новый лист
        ws = wb.active

        # Очистка листа
        ws.delete_cols(1, 100)
        ws.delete_rows(1, 3000)

        # Напишите заголовок в ячейку A1
        ws["A1"] = "datetime"
        ws["B1"] = "shop"
        ws["C1"] = "category"
        ws["D1"] = "product_name"
        ws["E1"] = "price"
        ws["F1"] = "volume"
        ws["G1"] = "price_real"
        ws["H1"] = "id"

        for row in rows:
            print(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7])

        for i in range(0, len(rows)):
            ws[f"A{i + 2}"] = rows[i][0]
            ws[f"B{i + 2}"], timezone_removal = str(rows[i][1]).split('+')
            ws[f"C{i + 2}"] = rows[i][2]
            ws[f"D{i + 2}"] = rows[i][3]
            ws[f"E{i + 2}"] = rows[i][4]
            ws[f"F{i + 2}"] = rows[i][5]
            ws[f"G{i + 2}"] = rows[i][6]
            ws[f"H{i + 2}"] = rows[i][7]

        # Сохраните файл Excel
        wb.save(f'../data/e_query_results/query_3-cheapest_products.xlsx')

    # Закрываем соединение
    cursor.close()
    connection.close()


if __name__ == '__main__':
    update_cheapest_products_xlsx()
