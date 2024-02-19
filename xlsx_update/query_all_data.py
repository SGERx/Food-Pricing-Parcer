import os
import sqlite3
import csv

import datetime as datetime
import openpyxl
import datetime
import pandas as pd
from openpyxl.utils import range_boundaries
from openpyxl.workbook import Workbook
import psycopg2


def update_all_data_xlsx():
    # Устанавливаем соединение с базой данных
    connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                                  port=5433)
    cursor = connection.cursor()

    # Выбираем все данные
    cursor.execute('''
    SELECT * FROM products;
    ''')

    # Выводим результаты
    # for result in cursor:
    #     print(result)

    rows = cursor.fetchall()
    print(rows)

    test_header = ['datetime', 'shop', 'category', 'product_name', 'price', 'volume', 'price_real', 'id']

    # with open(f'../data/e_query_results/test.csv', 'w', newline='') as file:
    #     writer = csv.writer(file, delimiter=' ', quotechar=',', quoting=csv.QUOTE_MINIMAL)
    #     writer.writerow(test_header)
    #     writer.writerows(rows)

    if not os.path.exists(f'../data/e_query_results/query_0-all_data.xlsx'):
        with open(f'../data/e_query_results/query_0-all_data.xlsx', 'w'):
            pass

    # OpenPyXL
    filepath = f'../data/e_query_results/query_0-all_data.xlsx'
    workbook_create = openpyxl.Workbook()
    workbook_create.save(filepath)

    wb = openpyxl.load_workbook(f'../data/e_query_results/query_0-all_data.xlsx')

    # Создайте новый лист
    ws = wb.active

    # Очистка листа
    ws.delete_cols(1, 100)
    ws.delete_rows(1, 3000)

    # Напишите заголовок в ячейку A1
    ws["A1"] = "datetime"
    ws["B1"] = "shop"
    ws["C1"] = "category"
    ws["D1"] = "product_name"
    ws["E1"] = "price"
    ws["F1"] = "volume"
    ws["G1"] = "price_real"
    ws["H1"] = "id"

    for row in rows:
        # print(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7])
        print(f" checking rows - {row[0]}, {row[1]}, {row[2]}, {row[3]}, {row[4]}, {row[5]}, {row[6]}, {row[7]}")

    for i in range(0, len(rows)):
        ws[f"A{i + 2}"] = rows[i][0]
        ws[f"B{i + 2}"], timezone_removal = str(rows[i][1]).split('+')
        ws[f"C{i + 2}"] = rows[i][2]
        ws[f"D{i + 2}"] = rows[i][3]
        ws[f"E{i + 2}"] = rows[i][4]
        ws[f"F{i + 2}"] = rows[i][5]
        ws[f"G{i + 2}"] = rows[i][6]
        ws[f"H{i + 2}"] = rows[i][7]

    # Сохраните файл Excel
    wb.save(f'../data/e_query_results/query_0-all_data.xlsx')

    # Закрываем соединение
    cursor.close()
    connection.close()


if __name__ == '__main__':
    update_all_data_xlsx()
