import os
import openpyxl
import psycopg2
from loguru import logger

def update_bill_price_xlsx():
    connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                                  port=5433)
    logger.info("Запуск функции {func}", func="update_bill_price_xlsx")

    logger.info("Создание курсора")
    cursor = connection.cursor()
    logger.info("Выполнение SQL-запроса")
    cursor.execute(f'''
    SELECT * 
    FROM products 
    WHERE datetime >= CURRENT_DATE - INTERVAL '3 days';
    ''')
    check_data = cursor.fetchall()
    if len(check_data) == 0:
        logger.info("Нет данных за последние три дня - произведите повторный парсинг")
    else:
        logger.info("Выполнение SQL-запроса")
        cursor.execute(f'''
        WITH min_prices AS (
        SELECT
        shop,
        category,
        MIN(price) as min_price
        FROM
        products WHERE datetime >= CURRENT_DATE - INTERVAL '3 days'
        GROUP BY
        shop, category
        )
        SELECT
        shop,
        SUM(min_price) as total_price
        FROM
        min_prices
        GROUP BY
        shop
        ORDER BY total_price ASC;
        ''')

        rows = cursor.fetchall()
        logger.info(rows)

        if not os.path.exists(f'../data/e_query_results/query_5-by_bill_price.xlsx'):
            logger.info("Создание файла с результатами")
            with open(f'../data/e_query_results/query_5-by_bill_price.xlsx', 'w'):
                pass

        filepath = f'../data/e_query_results/query_5-by_bill_price.xlsx'
        logger.info(f"Файл с результатами - {filepath}")
        workbook_create = openpyxl.Workbook()
        workbook_create.save(filepath)

        wb = openpyxl.load_workbook(f'../data/e_query_results/query_5-by_bill_price.xlsx')

        ws = wb.active

        ws.delete_cols(1, 100)
        ws.delete_rows(1, 3000)

        ws["A1"] = "shop"
        ws["B1"] = "sum_price"

        for row in rows:
            logger.info(row[0], row[1])

        for i in range(0, len(rows)):
            ws[f"A{i + 2}"] = rows[i][0]
            ws[f"B{i + 2}"] = rows[i][1]

        wb.save(f'../data/e_query_results/query_5-by_bill_price.xlsx')

    logger.info("Завершение функции {func}", func="update_bill_price_xlsx")


if __name__ == '__main__':
    logger.info("Запуск файла {file} через __main__", file="xlsx_bill_price.py")
    update_bill_price_xlsx()
    logger.info("Завершение файла {file} через __main__", file="xlsx_bill_price.py")
