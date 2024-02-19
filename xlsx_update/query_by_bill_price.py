import datetime
import os
import sqlite3
import csv

import openpyxl
import pandas as pd
import psycopg2

from datetime import timezone, timedelta, datetime

current_date = datetime.today().strftime('%Y-%m-%d')
current_date_timezoned = current_date + '+03'
dt = datetime.now(timezone.utc)
print(current_date)
print(current_date_timezoned)
print(dt)


def update_bill_price_xlsx():
    # Устанавливаем соединение с базой данных
    connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                                  port=5433)
    cursor = connection.cursor()

    # Проверяем данные за сегодняшнюю дату

    cursor.execute(f'''
    select * from products where datetime - '{dt}' < '1 day';
    ''')
    check_data = cursor.fetchall()
    print(len(check_data))
    print(check_data)
    if len(check_data) == 0:
        print('Нет данных за сегодняшнюю дату, используйте скрипт query_all_data.py для выгрузки всех данных и ручного '
              'просмотра')
    else:
        # Выбираем данные
        cursor.execute(f'''
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'auchan' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'globus' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'magnit' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'metro' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'miratorg' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'perekrestok' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'vkusvill' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        UNION
        SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
        FROM products
        WHERE id || ' ' || price_real IN (
            SELECT id || ' ' || MIN(price_real) 
            FROM products 
            WHERE shop = 'vprok' AND datetime - '{dt}' < INTERVAL '1 day'
            GROUP BY id
        )
        GROUP BY shop
        ORDER BY total_price ASC;
        ''')

        rows = cursor.fetchall()
        print(rows)

        test_header = ['sum_price_real', 'sum_price', 'shop']

        # with open(f'../data/e_query_results/by_price.csv', 'w', newline='') as file:
        #     writer = csv.writer(file)
        #     writer.writerow(test_header)
        #     writer.writerows(rows)

        if not os.path.exists(f'../data/e_query_results/query_1-by_bill_price.xlsx'):
            with open(f'../data/e_query_results/query_1-by_bill_price.xlsx', 'w'):
                pass

        # OpenPyXL
        filepath = f'../data/e_query_results/query_1-by_bill_price.xlsx'
        workbook_create = openpyxl.Workbook()
        workbook_create.save(filepath)

        wb = openpyxl.load_workbook(f'../data/e_query_results/query_1-by_bill_price.xlsx')

        # Создайте новый лист
        ws = wb.active

        # Очистка листа
        ws.delete_cols(1, 100)
        ws.delete_rows(1, 3000)

        # Напишите заголовок в ячейку A1
        ws["A1"] = "sum_price_real"
        ws["B1"] = "sum_price"
        ws["C1"] = "shop"

        for row in rows:
            print(row[0], row[1], row[2])

        for i in range(0, len(rows)):
            ws[f"A{i + 2}"] = rows[i][0]
            ws[f"B{i + 2}"] = rows[i][1]
            ws[f"C{i + 2}"] = rows[i][2]

        # Сохраните файл Excel
        wb.save(f'../data/e_query_results/query_1-by_bill_price.xlsx')

    def flask_bill_price():
        cursor.execute(f'''
        select * from products where datetime - '{dt}' < '1 day';
        ''')
        check_data = cursor.fetchall()
        print(len(check_data))
        print(check_data)
        if len(check_data) == 0:
            print(
                'Нет данных за сегодняшнюю дату, используйте скрипт query_all_data.py для выгрузки всех данных и ручного '
                'просмотра')
        else:
            # Выбираем данные
            cursor.execute(f'''
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'auchan' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'globus' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'magnit' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'metro' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'miratorg' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'perekrestok' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'vkusvill' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            UNION
            SELECT SUM(price_real) AS total_price_real, SUM(price) AS total_price, shop
            FROM products
            WHERE id || ' ' || price_real IN (
                SELECT id || ' ' || MIN(price_real) 
                FROM products 
                WHERE shop = 'vprok' AND datetime - '{dt}' < INTERVAL '1 day'
                GROUP BY id
            )
            GROUP BY shop
            ORDER BY total_price DESC;
            ''')

            data = cursor.fetchall()
            return data

    # Закрываем соединение
    cursor.close()
    connection.close()


if __name__ == '__main__':
    update_bill_price_xlsx()
