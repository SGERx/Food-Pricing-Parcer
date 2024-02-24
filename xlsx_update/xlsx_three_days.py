import os
import openpyxl
import psycopg2
from loguru import logger
connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                              port=5433)


def update_three_days_data_xlsx():
    logger.info("Запуск функции {func}", func="update_three_days_data_xlsx")
    logger.info("Создание курсора")
    cursor = connection.cursor()
    logger.info("Выполнение SQL-запроса")
    cursor.execute(f'''
    SELECT * 
    FROM products 
    WHERE datetime >= CURRENT_DATE - INTERVAL '3 days'
    ORDER BY shop, category, price, price_real;
    ''')
    rows = cursor.fetchall()
    if len(rows) == 0:
        logger.info("Нет данных за последние три дня - произведите повторный парсинг")
    logger.info(rows)

    if not os.path.exists(f'../data/e_query_results/query_1-three_days_data.xlsx'):
        logger.info("Создание файла с результатами")
        with open(f'../data/e_query_results/query_1-three_days_data.xlsx', 'w'):
            pass

    filepath = f'../data/e_query_results/query_1-three_days_data.xlsx'
    logger.info(f"Файл с результатами - {filepath}")
    workbook_create = openpyxl.Workbook()
    workbook_create.save(filepath)

    wb = openpyxl.load_workbook(f'../data/e_query_results/query_1-three_days_data.xlsx')

    ws = wb.active

    ws.delete_cols(1, 100)
    ws.delete_rows(1, 3000)

    ws["A1"] = "id"
    ws["B1"] = "datetime"
    ws["C1"] = "shop"
    ws["D1"] = "category"
    ws["E1"] = "product_name"
    ws["F1"] = "price"
    ws["G1"] = "volume"
    ws["H1"] = "price_real"

    for row in rows:
        logger.info(f" checking rows - {row[0]}, {row[1]}, {row[2]}, {row[3]}, {row[4]}, {row[5]}, {row[6]}, {row[7]}")

    for i in range(0, len(rows)):
        ws[f"A{i + 2}"] = rows[i][0]
        ws[f"B{i + 2}"], timezone_removal = str(rows[i][1]).split('.')
        ws[f"C{i + 2}"] = rows[i][2]
        ws[f"D{i + 2}"] = rows[i][3]
        ws[f"E{i + 2}"] = rows[i][4]
        ws[f"F{i + 2}"] = rows[i][5]
        ws[f"G{i + 2}"] = rows[i][6]
        ws[f"H{i + 2}"] = rows[i][7]

    logger.info("Запись данных")
    wb.save(f'../data/e_query_results/query_1-three_days_data.xlsx')
    logger.info("Завершение функции {func}", func="update_three_days_data_xlsx")


if __name__ == '__main__':
    logger.info("Запуск файла {file} через __main__", file="xlsx_three_days.py")
    update_three_days_data_xlsx()
    logger.info("Завершение файла {file} через __main__", file="xlsx_three_days.py")
