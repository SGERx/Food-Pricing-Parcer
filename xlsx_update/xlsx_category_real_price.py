import os
import openpyxl
import psycopg2
from loguru import logger
connection = psycopg2.connect(database="products_postgres", user="postgres", password="root", host="localhost",
                              port=5433)


def update_category__real_price_data_xlsx():
    logger.info("Запуск функции {func}", func="update_category__real_price_data_xlsx")
    logger.info("Создание курсора")
    cursor = connection.cursor()
    logger.info("Выполнение SQL-запроса")
    cursor.execute(f'''
    SELECT * 
    FROM products 
    WHERE datetime >= CURRENT_DATE - INTERVAL '3 days';
    ''')
    check_data = cursor.fetchall()
    if len(check_data) == 0:
        logger.info("Нет данных за последние три дня - произведите повторный парсинг")
    else:
        logger.info("Выполнение SQL-запроса")
        cursor.execute(f'''
        WITH ranked_products AS (
          SELECT
            shop,
            category,
            product_name,
            price_real,
            ROW_NUMBER() OVER (PARTITION BY shop, category ORDER BY price) as row_num
          FROM
            products
          WHERE
            datetime >= CURRENT_DATE - INTERVAL '3 days'
        )
        SELECT
          shop,
          category,
          product_name,
          price_real
        FROM
          ranked_products
        WHERE
          row_num = 1
        ORDER BY
          shop, category;
        ''')

        rows = cursor.fetchall()
        logger.info(rows)

        if not os.path.exists(f'../data/e_query_results/query_4-category_by_real_price.xlsx'):
            logger.info("Создание файла с результатами")
            with open(f'../data/e_query_results/query_4-category_by_real_price.xlsx', 'w'):
                pass

        filepath = f'../data/e_query_results/query_4-category_by_real_price.xlsx'
        logger.info(f"Файл с результатами - {filepath}")
        workbook_create = openpyxl.Workbook()
        workbook_create.save(filepath)

        wb = openpyxl.load_workbook(f'../data/e_query_results/query_4-category_by_real_price.xlsx')

        ws = wb.active

        ws.delete_cols(1, 100)
        ws.delete_rows(1, 3000)

        ws["A1"] = "shop"
        ws["B1"] = "category"
        ws["C1"] = "product_name"
        ws["D1"] = "price_real"

        for row in rows:
            logger.info(row[0], row[1], row[2], row[3])

        for i in range(0, len(rows)):
            ws[f"A{i + 2}"] = rows[i][0]
            ws[f"B{i + 2}"] = rows[i][1]
            ws[f"C{i + 2}"] = rows[i][2]
            ws[f"D{i + 2}"] = rows[i][3]

        wb.save(f'../data/e_query_results/query_4-category_by_real_price.xlsx')

    logger.info("Завершение функции {func}", func="update_category__real_price_data_xlsx")


if __name__ == '__main__':
    logger.info("Запуск файла {file} через __main__", file="xlsx_category_real_price.py")
    update_category__real_price_data_xlsx()
    logger.info("Завершение файла {file} через __main__", file="xlsx_category_real_price.py")
